import re
import json
import os
import time
from datetime import datetime
from urllib.parse import quote

import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

XLSX_FILE = "image_urls.xlsx"
HEADERS_ROW = ["date", "source", "keyword", "image_url"]

# DuckDuckGo 用ヘッダー
# /i.js が GET→403 になったため POST + XHR ヘッダーが必要
DDG_REGION = "jp-jp"
DDG_HEADERS_GET = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
    "Accept-Language": "ja,en-US;q=0.9",
    "Referer": "https://duckduckgo.com/",
}
DDG_HEADERS_POST = {
    **DDG_HEADERS_GET,
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Content-Type": "application/x-www-form-urlencoded",
    "X-Requested-With": "XMLHttpRequest",
}
# 後方互換のため旧名もエイリアスとして残す
DDG_HEADERS = DDG_HEADERS_GET


# ---------- 共通ヘルパー ----------

def excel_safe(value):
    """数式インジェクション対策: 数式として解釈されうる先頭文字を無害化する"""
    s = str(value)
    if s and s[0] in ("=", "+", "-", "@"):
        s = "'" + s
    return s


def load_seen(keyword):
    """既存のExcelから『同じキーワード』のURLを読み込み、重複チェック用のsetを返す"""
    seen = set()
    if not os.path.exists(XLSX_FILE):
        return seen
    wb = load_workbook(XLSX_FILE, read_only=True)
    ws = wb.active
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:  # ヘッダー行を飛ばす
            first = False
            continue
        if not row or len(row) < 4:
            continue
        kw, url = row[2], row[3]
        if url and kw in (keyword, excel_safe(keyword)):
            seen.add(url)
    wb.close()
    return seen


def _format_sheet(ws):
    """ヘッダーを太字＋枠固定し、列幅を内容に合わせて整える"""
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None),
                      default=0)
        ws.column_dimensions[letter].width = min(max_len + 2, 80)


def save_urls(source_label, keyword, urls):
    """取得したURLをExcelに追記する（image_url列はクリックできるリンクにする）"""
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "images"
        ws.append(HEADERS_ROW)

    today = datetime.now().strftime("%Y-%m-%d")
    for url in urls:
        ws.append([today, source_label, excel_safe(keyword), url])
        link_cell = ws.cell(row=ws.max_row, column=4)
        link_cell.hyperlink = url
        link_cell.font = Font(color="0563C1", underline="single")

    _format_sheet(ws)
    wb.save(XLSX_FILE)


# ---------- 取得元ごとの処理 ----------

_CHROME_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
)


def _make_driver():
    """ヘッドレスChromeドライバーを生成。自動化フラグを隠してbot検知を回避する"""
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1600,1200")
    # navigator.webdriver / HeadlessChrome UA の痕跡を消す
    options.add_argument(f"--user-agent={_CHROME_UA}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(options=options)
    # CDP 経由で UA を確実に上書き（--headless の痕跡を消す）
    driver.execute_cdp_cmd("Network.setUserAgentOverride", {"userAgent": _CHROME_UA})
    return driver


def scrape_bing(keyword, limit, seen):
    """Bing画像検索。.dgControl_list 内の a.iusc から murl を取る"""
    urls = []
    search_url = f"https://www.bing.com/images/search?q={quote(keyword)}"
    driver = _make_driver()
    try:
        driver.get(search_url)
        # 検索結果エリアに a.iusc が10件以上現れるまで最大15秒待つ
        try:
            WebDriverWait(driver, 15).until(
                lambda d: len(d.find_elements(
                    By.CSS_SELECTOR, ".dgControl_list a.iusc")) >= 10
            )
        except TimeoutException:
            pass  # タイムアウトしても取れた分で続行

        last_count = 0
        stale = 0
        while len(urls) < limit and stale < 5:
            # .dgControl_list に絞って広告・関連コンテナを除外する
            for a in driver.find_elements(By.CSS_SELECTOR, ".dgControl_list a.iusc"):
                m = a.get_attribute("m")
                if not m:
                    continue
                try:
                    info = json.loads(m)
                except json.JSONDecodeError:
                    continue
                murl = info.get("murl")
                if not murl or not murl.startswith("http"):
                    continue
                if murl in seen:
                    continue
                seen.add(murl)
                urls.append(murl)
                if len(urls) >= limit:
                    break

            print(f"取得済み: {len(urls)}件")
            if len(urls) >= limit:
                break
            stale = stale + 1 if len(urls) == last_count else 0
            last_count = len(urls)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
    finally:
        driver.quit()
    return urls


# i.pinimg.com/<size_token>/<path> の size_token 部分にマッチ
# 例: 236x / 474x / 564x / 736x / originals
_PINIMG_SIZE_RE = re.compile(r"(?<=https://i\.pinimg\.com/)[^/]+(?=/)")


_PINIMG_HEAD_HEADERS = {
    "User-Agent": _CHROME_UA,
    "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
    "Referer": "https://www.pinterest.com/",
    "Sec-Fetch-Dest": "image",
    "Sec-Fetch-Mode": "no-cors",
    "Sec-Fetch-Site": "cross-site",
}


def _pinterest_best_url(src):
    """i.pinimg.com の URL を originals に昇格させる。
    HEAD リクエストで 200 以外なら 736x（大きめ）にフォールバックする。
    """
    originals = _PINIMG_SIZE_RE.sub("originals", src)
    if originals == src:
        # マッチしない（既に originals / パターン外）はそのまま返す
        return src
    try:
        r = requests.head(
            originals,
            timeout=6,
            allow_redirects=True,
            headers=_PINIMG_HEAD_HEADERS,
        )
        if r.status_code == 200:
            return originals
    except requests.RequestException:
        pass
    # originals が 200 でなければ 736x へフォールバック
    return _PINIMG_SIZE_RE.sub("736x", src)


def _dismiss_modal(driver):
    """Pinterestのモーダルを安全に閉じる。閉じるボタンだけを狙い、無ければ何もしない"""
    selectors = [
        "//button[@aria-label='閉じる']",
        "//button[@aria-label='Close']",
        "//div[@role='dialog']//button[contains(@aria-label, '閉じ')]",
        "//div[@role='dialog']//button[contains(@aria-label, 'Close')]",
    ]
    for xp in selectors:
        try:
            driver.find_element(By.XPATH, xp).click()
            time.sleep(1.5)
            return
        except (NoSuchElementException, WebDriverException):
            continue


def scrape_pinterest(keyword, limit, seen):
    """Pinterest検索。i.pinimg.com の画像URLを収集する"""
    urls = []
    search_url = f"https://www.pinterest.com/search/pins/?q={quote(keyword)}"
    driver = _make_driver()
    try:
        driver.get(search_url)
        time.sleep(5)
        _dismiss_modal(driver)

        last_count = 0
        stale = 0
        while len(urls) < limit and stale < 5:
            for img in driver.find_elements(By.TAG_NAME, "img"):
                try:
                    src = img.get_attribute("src")
                except StaleElementReferenceException:
                    continue
                if not src or not src.startswith("http"):
                    continue
                if "i.pinimg.com" not in src:
                    continue
                if "/60x60/" in src:  # 極小サムネイルは除外
                    continue
                # originals に昇格（404 なら 736x にフォールバック）
                best = _pinterest_best_url(src)
                if best in seen:
                    continue
                seen.add(best)
                urls.append(best)
                if len(urls) >= limit:
                    break

            print(f"取得済み: {len(urls)}件")
            if len(urls) >= limit:
                break
            stale = stale + 1 if len(urls) == last_count else 0
            last_count = len(urls)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
    finally:
        driver.quit()
    return urls


def _get_vqd(session, keyword):
    """DuckDuckGo画像検索ページから vqd トークンを取り出す"""
    resp = session.get(
        "https://duckduckgo.com/",
        params={"q": keyword, "iax": "images", "ia": "images"},
        headers=DDG_HEADERS_GET,
        timeout=15,
    )
    resp.raise_for_status()
    m = re.search(r'vqd=["\']?([\d-]+)', resp.text)
    return m.group(1) if m else None


def scrape_ddg(keyword, limit, seen):
    """DuckDuckGo画像検索。POST /i.js のJSONから元画像URL(image)を取る（1バッチ）"""
    urls = []
    session = requests.Session()

    try:
        vqd = _get_vqd(session, keyword)
    except requests.RequestException as e:
        print(f"検索ページへの接続に失敗しました: {e}")
        return urls
    if not vqd:
        print("vqdトークンを取得できませんでした。"
              "DuckDuckGo側の仕様変更か、一時的なブロックの可能性があります。")
        return urls

    # /i.js は 2024年以降 GET→403 のため POST で送信する
    body = {
        "l": DDG_REGION,
        "o": "json",
        "q": keyword,
        "vqd": vqd,
        "f": ",,,",
        "p": "1",
    }
    try:
        resp = session.post(
            "https://duckduckgo.com/i.js",
            headers=DDG_HEADERS_POST,
            data=body,
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as e:
        print(f"画像データの取得に失敗しました: {e}")
        return urls
    except ValueError:
        print("画像データをJSONとして読み取れませんでした。"
              "少し時間をおいて再実行してください（レート制限の可能性があります）。")
        return urls

    for item in data.get("results", []):
        img = item.get("image")  # image = 元画像のURL
        if not img or not img.startswith("http"):
            continue
        if img in seen:
            continue
        seen.add(img)
        urls.append(img)
        if len(urls) >= limit:
            break

    print(f"取得済み: {len(urls)}件")
    return urls


def scrape_unsplash(keyword, limit, seen, api_key=""):
    """Unsplash API で画像検索し、regular サイズの URL を返す"""
    urls = []
    page = 1
    per_page = min(limit, 30)  # Unsplash API の1リクエスト上限は30

    while len(urls) < limit:
        try:
            resp = requests.get(
                "https://api.unsplash.com/search/photos",
                params={"query": keyword, "per_page": per_page, "page": page},
                headers={"Authorization": f"Client-ID {api_key}"},
                timeout=15,
            )
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as e:
            print(f"Unsplash APIへの接続に失敗しました: {e}")
            break

        results = data.get("results", [])
        if not results:
            break

        for photo in results:
            url = photo.get("urls", {}).get("regular")
            if not url or url in seen:
                continue
            seen.add(url)
            urls.append(url)
            if len(urls) >= limit:
                break

        print(f"取得済み: {len(urls)}件")

        if len(results) < per_page:  # 最終ページ
            break
        page += 1

    return urls


def scrape_pexels(keyword, limit, seen, api_key=""):
    """Pexels API で画像検索し、large サイズの URL を返す"""
    urls = []
    page = 1
    per_page = min(limit, 80)  # Pexels API の1リクエスト上限は80

    while len(urls) < limit:
        try:
            resp = requests.get(
                "https://api.pexels.com/v1/search",
                params={"query": keyword, "per_page": per_page, "page": page},
                headers={"Authorization": api_key},
                timeout=15,
            )
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as e:
            print(f"Pexels APIへの接続に失敗しました: {e}")
            break

        photos = data.get("photos", [])
        if not photos:
            break

        for photo in photos:
            url = photo.get("src", {}).get("large")
            if not url or url in seen:
                continue
            seen.add(url)
            urls.append(url)
            if len(urls) >= limit:
                break

        print(f"取得済み: {len(urls)}件")

        if not data.get("next_page"):  # 最終ページ
            break
        page += 1

    return urls


# ---------- メイン ----------

SOURCES = {
    "1": ("Bing", "bing-original", scrape_bing),
    "2": ("Pinterest", "pinterest", scrape_pinterest),
    "3": ("DuckDuckGo", "duckduckgo", scrape_ddg),
    "4": ("Unsplash", "unsplash", scrape_unsplash),
    "5": ("Pexels", "pexels", scrape_pexels),
}


def main():
    keyword = input("検索キーワードを入力してください: ")

    # 取得件数（1以上の整数になるまで聞き直す）
    while True:
        raw = input("取得件数を入力してください: ")
        try:
            limit = int(raw)
            if limit > 0:
                break
        except ValueError:
            pass
        print("1以上の整数を入力してください。")

    # 取得元の選択
    print("取得元を選んでください:")
    print("  1: Bing")
    print("  2: Pinterest")
    print("  3: DuckDuckGo")
    while True:
        choice = input("番号を入力 (1/2/3): ").strip()
        if choice in SOURCES:
            break
        print("1, 2, 3 のいずれかを入力してください。")

    name, label, scraper = SOURCES[choice]

    seen = load_seen(keyword)
    if seen:
        print(f"同じキーワードの既存URL {len(seen)}件を重複チェック対象に読み込みました。")

    print(f"{name} から「{keyword}」の画像URLを取得します...")
    urls = scraper(keyword, limit, seen)[:limit]

    save_urls(label, keyword, urls)
    print(f"{len(urls)}件の画像URLを {XLSX_FILE} に保存しました。（取得元: {name}）")


if __name__ == "__main__":
    main()
